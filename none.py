import cv2
import os
import numpy as np
import openpyxl
import subprocess
from datetime import datetime

def load_images_from_folder(folder, target_size=(100, 100)):
    images = []
    labels = []
    label_names = {}  # store label names
    for label, person_folder in enumerate(os.listdir(folder)):
        label_names[label] = person_folder  # map label to person name
        for filename in os.listdir(os.path.join(folder, person_folder)):
            img_path = os.path.join(folder, person_folder, filename)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if img is not None:
                img_resized = cv2.resize(img, target_size)
                images.append(img_resized)
                labels.append(label)
    return images, np.array(labels), label_names

#  dataset
dataset_folder = 'dataset'
images, labels, label_names = load_images_from_folder(dataset_folder)

# train  model
model = cv2.face.EigenFaceRecognizer_create()
model.train(images, labels)

# camera
camera = cv2.VideoCapture(0)   # 0(default camera)

excel_file = 'face_recognition_entries.xlsx'
if not os.path.exists(excel_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ID', 'Name', 'University', 'Date', 'Time'])
    wb.save(excel_file)

# existing data from the excel file
wb = openpyxl.load_workbook(excel_file)
ws = wb.active
existing_person_data = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3)]

# track person data enter( current run)
entered_data = set()

while True:
    ret, frame = camera.read()
    if not ret:
        print("Error: Failed to capture image from camera.")
        break

    #  frame to grayscale for face detection
    gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # detect faces 

    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
    faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))


    # no faces 
    if len(faces) == 0:
        print("No face detected. Running face.py to register a new face.")
        subprocess.run(['python', 'face.py'])
        break

    # recognize faces
    for (x, y, w, h) in faces:

        #  face details
        face_roi = gray_frame[y:y+h, x:x+w]
        
        # resize  face region match the model input size
        face_roi_resized = cv2.resize(face_roi, (100, 100))

        # predict  face
        label, confidence = model.predict(face_roi_resized)

        # recognized face and confidence level
        if confidence < 3000:
            person_folder = label_names.get(label, "Unknown")
            person_id, person_name, university_name = person_folder.split('_')

            # details in a box on the right side
            text = f"ID: {person_id}\nName: {person_name}\nUniversity: {university_name}"
            box_x = x + w + 10
            box_y = y
            y0, dy = box_y, 30

            #draw text box
            box_height = dy * 3
            box_width = max([cv2.getTextSize(line, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)[0][0] for line in text.split('\n')]) + 10
            cv2.rectangle(frame, (box_x, box_y), (box_x + box_width, box_y + box_height), (0, 255, 0), 2)

            for i, line in enumerate(text.split('\n')):
                y = y0 + i * dy + 20
                cv2.putText(frame, line, (box_x + 5, y), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            if (person_id, person_name, university_name) not in entered_data:
                if (person_id, person_name, university_name) not in existing_person_data:
                    # add person data file
                    current_datetime = datetime.now()
                    date_str = current_datetime.strftime("%Y-%m-%d")
                    time_str = current_datetime.strftime("%H:%M:%S")
                    ws.append([person_id, person_name, university_name, date_str, time_str])
                    existing_person_data.append((person_id, person_name, university_name))
                    wb.save(excel_file)
                    print(f"Added {person_name} from {university_name} to the Excel file for the first time on {date_str} at {time_str}.")
                print(f"Alert: {person_name}'s face detected again!")
                entered_data.add((person_id, person_name, university_name))

        else:
            cv2.putText(frame, 'Unknown', (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

        cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)

    cv2.imshow('Face Recognition', frame)

    key = cv2.waitKey(1)
    if key == ord('q'):  # q  to exit
        break

camera.release()
cv2.destroyAllWindows()
